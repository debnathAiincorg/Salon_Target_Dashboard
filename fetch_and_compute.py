import os
import json
import base64
import io
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
import requests
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# Constants
WEEKLY_TARGET = 15000
OUTPUT_FILE = "dashboard-data.json"

# Azure AD configuration
AZURE_TENANT_ID = os.getenv("AZURE_TENANT_ID")
AZURE_CLIENT_ID = os.getenv("AZURE_CLIENT_ID")
AZURE_CLIENT_SECRET = os.getenv("AZURE_CLIENT_SECRET")

# SharePoint share link
SHAREPOINT_SHARE_LINK = "https://cloudaiorg.sharepoint.com/:x:/r/sites/Accounts2/_layouts/15/Doc.aspx?sourcedoc=%7B11D86568-2055-430C-BC22-014479DF6ACD%7D&file=Daily_Invoice.xlsx&fromShare=true&action=default&mobileredirect=true"

def get_access_token():
    """Authenticate via Azure AD client credentials flow and return access token."""
    if not AZURE_TENANT_ID or not AZURE_CLIENT_ID or not AZURE_CLIENT_SECRET:
        raise ValueError(
            "Azure AD credentials not configured. "
            "Please set AZURE_TENANT_ID, AZURE_CLIENT_ID, and AZURE_CLIENT_SECRET in .env"
        )

    token_url = f"https://login.microsoftonline.com/{AZURE_TENANT_ID}/oauth2/v2.0/token"
    payload = {
        "grant_type": "client_credentials",
        "client_id": AZURE_CLIENT_ID,
        "client_secret": AZURE_CLIENT_SECRET,
        "scope": "https://graph.microsoft.com/.default",
    }

    try:
        response = requests.post(token_url, data=payload, timeout=10)
        response.raise_for_status()
        return response.json()["access_token"]
    except requests.exceptions.RequestException as e:
        raise RuntimeError(f"Failed to authenticate with Azure AD: {e}")
    except (KeyError, ValueError) as e:
        raise RuntimeError(f"Unexpected response from Azure AD token endpoint: {e}")

def resolve_sharepoint_link_to_driveitem(access_token, share_link):
    """Resolve SharePoint share link to driveItem using Graph API /shares endpoint."""
    # Base64-url encode the share link: prefix "u!", base64-url encode, strip padding
    share_link_bytes = share_link.encode('utf-8')
    encoded = base64.urlsafe_b64encode(share_link_bytes).decode('utf-8').rstrip('=')
    share_id = f"u!{encoded}"

    shares_url = f"https://graph.microsoft.com/v1.0/shares/{share_id}/driveItem"
    headers = {"Authorization": f"Bearer {access_token}"}

    try:
        response = requests.get(shares_url, headers=headers, timeout=10)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        raise RuntimeError(f"Failed to resolve SharePoint share link: {e}")
    except (KeyError, ValueError) as e:
        raise RuntimeError(f"Unexpected response from Graph API shares endpoint: {e}")

def download_file_from_sharepoint(access_token, drive_item):
    """Download file content from SharePoint driveItem."""
    # Use @microsoft.graph.downloadUrl to get direct download link
    if "@microsoft.graph.downloadUrl" not in drive_item:
        raise RuntimeError(
            "Share link resolved but downloadUrl not available. "
            "Check that the file exists and is accessible."
        )

    download_url = drive_item["@microsoft.graph.downloadUrl"]

    try:
        response = requests.get(download_url, timeout=30)
        response.raise_for_status()
        return response.content
    except requests.exceptions.RequestException as e:
        raise RuntimeError(f"Failed to download file from SharePoint: {e}")

def fetch_excel_from_sharepoint():
    """Fetch Daily_Invoice.xlsx from SharePoint via Graph API and return openpyxl Workbook."""
    # Step 1: Authenticate
    access_token = get_access_token()

    # Step 2: Resolve share link to driveItem
    drive_item = resolve_sharepoint_link_to_driveitem(access_token, SHAREPOINT_SHARE_LINK)

    # Step 3: Download file content
    file_content = download_file_from_sharepoint(access_token, drive_item)

    # Step 4: Load into openpyxl from bytes
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        return wb
    except Exception as e:
        raise RuntimeError(f"Failed to load Excel file into openpyxl: {e}")

def parse_excel(ws):
    """Parse invoice_transactions sheet and aggregate NET_AMOUNT by DATE.

    Reads raw transaction data from invoice_transactions sheet.
    Multiple rows per date are summed into one daily_sales record.

    Args:
        ws: openpyxl worksheet object (invoice_transactions sheet)

    Returns:
        List of dicts with keys 'date' (datetime.date) and 'daily_sales' (float),
        sorted by date ascending. Skips rows where DATE (column C) or
        NET_AMOUNT (column D) is None or unparseable.
    """
    # Aggregate by date: {date_obj: total_net_amount}
    daily_totals = {}

    # Iterate rows starting from row 2 (row 1 is header)
    # Column C (index 2) = DATE, Column D (index 3) = NET_AMOUNT
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_value = row[2] if len(row) > 2 else None
        net_amount_value = row[3] if len(row) > 3 else None

        # Skip if date or net_amount is None
        if date_value is None or net_amount_value is None:
            continue

        # Parse date
        if isinstance(date_value, datetime):
            date_obj = date_value.date()
        else:
            try:
                date_obj = datetime.strptime(str(date_value), "%Y-%m-%d").date()
            except:
                print(f"WARNING: Skipping row with unparseable date: {date_value}")
                continue

        # Parse net amount
        try:
            net_amount = float(net_amount_value)
        except:
            print(f"WARNING: Skipping row with unparseable NET_AMOUNT: {net_amount_value} for date {date_obj}")
            continue

        # Aggregate: sum all net amounts for this date
        if date_obj not in daily_totals:
            daily_totals[date_obj] = 0
        daily_totals[date_obj] += net_amount

    # Convert to daily_records format and sort
    daily_records = [
        {"date": date_obj, "daily_sales": total}
        for date_obj, total in daily_totals.items()
    ]
    daily_records.sort(key=lambda x: x["date"])

    return daily_records

def compute_weekly_summary(daily_records):
    """Compute weekly sales summaries from daily records.

    Week = Monday (weekday 0) to Sunday (weekday 6).
    For each week, compute total_sales and target_pending.

    Args:
        daily_records: List of dicts with 'date' and 'daily_sales'

    Returns:
        Dict keyed by week_start_date (datetime.date).
        Each value: {"week_start_date": date, "week_end_date": date,
                     "total_sales": float, "target_pending": float}
    """
    if not daily_records:
        return {}

    # Group records by week
    weeks = {}
    for record in daily_records:
        date = record["date"]
        week_start = date - timedelta(days=date.weekday())  # Monday of that week
        week_end = week_start + timedelta(days=6)  # Sunday of that week

        if week_start not in weeks:
            weeks[week_start] = {
                "week_start_date": week_start,
                "week_end_date": week_end,
                "total_sales": 0,
            }

        weeks[week_start]["total_sales"] += record["daily_sales"]

    # Compute target_pending for each week
    for week_start, week_data in weeks.items():
        total_sales = week_data["total_sales"]
        if total_sales >= WEEKLY_TARGET:
            week_data["target_pending"] = 0
        else:
            week_data["target_pending"] = WEEKLY_TARGET - total_sales

    return weeks

def get_current_month_sales(daily_records):
    """Filter daily records to current month only.

    Args:
        daily_records: List of dicts with 'date' and 'daily_sales'

    Returns:
        Filtered list, sorted by date ascending
    """
    today = datetime.now().date()
    current_month = [
        record for record in daily_records
        if record["date"].year == today.year and record["date"].month == today.month
    ]
    return current_month

def compute_kpi_metrics(daily_records, weekly_summary):
    """Compute all KPI metrics for the dashboard.

    Args:
        daily_records: List of dicts with 'date' and 'daily_sales'
        weekly_summary: Dict keyed by week_start_date

    Returns:
        Dict with KPI metrics:
        - current_date_label: str (e.g., "23 May")
        - current_week_sales: float
        - current_week_balance: float (target_pending)
        - remaining_days: int
        - previous_week_sales: float
        - monthly_sales: float
    """
    today = datetime.now().date()

    # 1. Current Date Label
    current_date_label = f"{today.day} {today.strftime('%B')}"

    # 2. Current Week Sales and Balance
    # Find most recent week_start_date <= today
    week_starts = sorted([ws for ws in weekly_summary.keys() if ws <= today], reverse=True)
    if week_starts:
        current_week_start = week_starts[0]
        current_week_sales = weekly_summary[current_week_start]["total_sales"]
        current_week_balance = weekly_summary[current_week_start]["target_pending"]
    else:
        current_week_sales = 0
        current_week_balance = 0

    # 3. Remaining Days
    remaining_days = 7 - today.weekday()

    # 4. Previous Week Sales
    # Find second-most-recent week
    if len(week_starts) >= 2:
        previous_week_start = week_starts[1]
        previous_week_sales = weekly_summary[previous_week_start]["total_sales"]
    else:
        previous_week_sales = 0

    # 5. Monthly Sales
    current_month_sales = get_current_month_sales(daily_records)
    monthly_sales = sum(record["daily_sales"] for record in current_month_sales)

    return {
        "current_date_label": current_date_label,
        "current_week_sales": current_week_sales,
        "current_week_balance": current_week_balance,
        "remaining_days": remaining_days,
        "previous_week_sales": previous_week_sales,
        "monthly_sales": monthly_sales,
        "current_month_sales": current_month_sales,  # For chart data
    }

def generate_chart_data(current_month_sales):
    """Generate chart data from current month sales records.

    Args:
        current_month_sales: List of dicts with 'date' and 'daily_sales'

    Returns:
        List of dicts with 'date' (formatted string) and 'daily_sales' (int)
    """
    chart_data = []
    for record in current_month_sales:
        date_obj = record["date"]
        # Optional: use "%b" instead of "%B" for abbreviated month names (e.g., "Jun 24" vs "June 24")
        # Abbreviated format may reduce label crowding with a full month of data
        month_name = date_obj.strftime("%B")
        day = date_obj.day
        date_str = f"{month_name} {day}"

        chart_data.append({
            "date": date_str,
            "daily_sales": int(record["daily_sales"])
        })

    return chart_data

def write_dashboard_json(kpi, chart_data, output_file):
    """Write dashboard data to JSON file.

    Args:
        kpi: Dict with KPI metrics (current_date_label, current_week_sales, etc.)
        chart_data: List of dicts with chart data
        output_file: Path to output JSON file
    """
    dashboard_data = {
        "current_date_label": kpi["current_date_label"],
        "current_week_sales": int(kpi["current_week_sales"]),
        "current_week_balance": int(kpi["current_week_balance"]),
        "remaining_days": kpi["remaining_days"],
        "previous_week_sales": int(kpi["previous_week_sales"]),
        "monthly_sales": int(kpi["monthly_sales"]),
        "daily_chart_data": chart_data,
    }

    with open(output_file, 'w') as f:
        json.dump(dashboard_data, f, indent=2)

    print(f"Dashboard data written to {output_file}")

def main():
    """Main entry point for the fetch and compute script."""
    try:
        # Step 1: Fetch Excel from SharePoint via Graph API
        print("Fetching Daily_Invoice.xlsx from SharePoint...")
        wb = fetch_excel_from_sharepoint()

        # Step 2: Validate sheet exists
        if "invoice_transactions" not in wb.sheetnames:
            print("ERROR: Sheet 'invoice_transactions' not found in Excel file")
            exit(1)

        ws = wb["invoice_transactions"]
        print("[OK] Successfully fetched and opened Daily_Invoice.xlsx")

        # Step 3: Parse and aggregate invoice_transactions by date
        daily_records = parse_excel(ws)

        if not daily_records:
            print("WARNING: No valid data rows found in Excel file")
            daily_records = []
        else:
            print(f"Parsed {len(daily_records)} daily records from Excel")
            print(f"Date range: {daily_records[0]['date']} to {daily_records[-1]['date']}")

        # Step 4: Compute weekly summaries
        weekly_summary = compute_weekly_summary(daily_records)
        print(f"Computed {len(weekly_summary)} weekly summaries")

        # Step 5: Compute KPI metrics
        kpi = compute_kpi_metrics(daily_records, weekly_summary)
        print(f"Current Date: {kpi['current_date_label']}")
        print(f"Current Week Sales: {kpi['current_week_sales']}")
        print(f"Monthly Sales: {kpi['monthly_sales']}")

        # Step 6: Generate chart data
        chart_data = generate_chart_data(kpi["current_month_sales"])

        # Step 7: Write output JSON
        write_dashboard_json(kpi, chart_data, OUTPUT_FILE)

        print("SUCCESS: Dashboard data computed and written")

        wb.close()

    except RuntimeError as e:
        print(f"ERROR: {e}")
        exit(1)
    except Exception as e:
        print(f"ERROR: Unexpected error: {e}")
        exit(1)

if __name__ == "__main__":
    main()
