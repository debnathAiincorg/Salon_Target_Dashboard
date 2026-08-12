import os
import json
import base64
import io
from datetime import datetime
from zoneinfo import ZoneInfo
import openpyxl
import requests

# Load .env file if available (local development)
# In CI/GitHub Actions, env vars are injected directly; dotenv is not needed
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# Constants
OUTPUT_FILE = "monthly-dashboard-data.json"

# Azure AD configuration
AZURE_TENANT_ID = os.getenv("AZURE_TENANT_ID")
AZURE_CLIENT_ID = os.getenv("AZURE_CLIENT_ID")
AZURE_CLIENT_SECRET = os.getenv("AZURE_CLIENT_SECRET")

# SharePoint configuration
SHAREPOINT_SHARE_LINK = os.getenv("SHAREPOINT_SHARE_LINK")

def get_access_token():
    """Authenticate via Azure AD client credentials flow and return access token."""
    if not AZURE_TENANT_ID or not AZURE_CLIENT_ID or not AZURE_CLIENT_SECRET:
        raise ValueError(
            "Azure AD credentials not configured. "
            "Please set AZURE_TENANT_ID, AZURE_CLIENT_ID, and AZURE_CLIENT_SECRET in .env"
        )

    token_url = f"https://login.microsoftonline.com/{AZURE_TENANT_ID}/oauth2/v2.0/token"
    payload = {
        "grant_type": "client_credentials",
        "client_id": AZURE_CLIENT_ID,
        "client_secret": AZURE_CLIENT_SECRET,
        "scope": "https://graph.microsoft.com/.default",
    }

    try:
        response = requests.post(token_url, data=payload, timeout=10)
        response.raise_for_status()
        return response.json()["access_token"]
    except requests.exceptions.RequestException as e:
        raise RuntimeError(f"Failed to authenticate with Azure AD: {e}")
    except (KeyError, ValueError) as e:
        raise RuntimeError(f"Unexpected response from Azure AD token endpoint: {e}")

def resolve_sharepoint_link_to_driveitem(access_token, share_link):
    """Resolve SharePoint share link to driveItem using Graph API /shares endpoint."""
    # Base64-url encode the share link: prefix "u!", base64-url encode, strip padding
    share_link_bytes = share_link.encode('utf-8')
    encoded = base64.urlsafe_b64encode(share_link_bytes).decode('utf-8').rstrip('=')
    share_id = f"u!{encoded}"

    shares_url = f"https://graph.microsoft.com/v1.0/shares/{share_id}/driveItem"
    headers = {"Authorization": f"Bearer {access_token}"}

    try:
        response = requests.get(shares_url, headers=headers, timeout=10)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        raise RuntimeError(f"Failed to resolve SharePoint share link: {e}")
    except (KeyError, ValueError) as e:
        raise RuntimeError(f"Unexpected response from Graph API shares endpoint: {e}")

def download_file_from_sharepoint(access_token, drive_item):
    """Download file content from SharePoint driveItem."""
    # Use @microsoft.graph.downloadUrl to get direct download link
    if "@microsoft.graph.downloadUrl" not in drive_item:
        raise RuntimeError(
            "Share link resolved but downloadUrl not available. "
            "Check that the file exists and is accessible."
        )

    download_url = drive_item["@microsoft.graph.downloadUrl"]

    try:
        response = requests.get(download_url, timeout=30)
        response.raise_for_status()
        return response.content
    except requests.exceptions.RequestException as e:
        raise RuntimeError(f"Failed to download file from SharePoint: {e}")

def fetch_excel_from_sharepoint():
    """Fetch Daily_Invoice.xlsx from SharePoint via Graph API and return openpyxl Workbook."""
    # Validate configuration
    if not SHAREPOINT_SHARE_LINK:
        raise ValueError(
            "SharePoint share link not configured. "
            "Please set SHAREPOINT_SHARE_LINK in .env"
        )

    # Step 1: Authenticate
    access_token = get_access_token()

    # Step 2: Resolve share link to driveItem
    drive_item = resolve_sharepoint_link_to_driveitem(access_token, SHAREPOINT_SHARE_LINK)

    # Step 3: Download file content
    file_content = download_file_from_sharepoint(access_token, drive_item)

    # Step 4: Load into openpyxl from bytes
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        return wb
    except Exception as e:
        raise RuntimeError(f"Failed to load Excel file into openpyxl: {e}")

def parse_excel(ws):
    """Parse invoice_transactions sheet and aggregate NET_AMOUNT by DATE.

    Reads raw transaction data from invoice_transactions sheet.
    Multiple rows per date are summed into one daily_sales record.

    Args:
        ws: openpyxl worksheet object (invoice_transactions sheet)

    Returns:
        List of dicts with keys 'date' (datetime.date) and 'daily_sales' (float),
        sorted by date ascending. Skips rows where DATE (column C) or
        NET_AMOUNT (column D) is None or unparseable.
    """
    # Aggregate by date: {date_obj: total_net_amount}
    daily_totals = {}

    # Iterate rows starting from row 2 (row 1 is header)
    # Column C (index 2) = DATE, Column D (index 3) = NET_AMOUNT
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_value = row[2] if len(row) > 2 else None
        net_amount_value = row[3] if len(row) > 3 else None

        # Skip if date or net_amount is None
        if date_value is None or net_amount_value is None:
            continue

        # Parse date
        if isinstance(date_value, datetime):
            date_obj = date_value.date()
        else:
            try:
                date_obj = datetime.strptime(str(date_value), "%Y-%m-%d").date()
            except:
                print(f"WARNING: Skipping row with unparseable date: {date_value}")
                continue

        # Parse net amount
        try:
            net_amount = float(net_amount_value)
        except:
            print(f"WARNING: Skipping row with unparseable NET_AMOUNT: {net_amount_value} for date {date_obj}")
            continue

        # Aggregate: sum all net amounts for this date
        if date_obj not in daily_totals:
            daily_totals[date_obj] = 0
        daily_totals[date_obj] += net_amount

    # Convert to daily_records format and sort
    daily_records = [
        {"date": date_obj, "daily_sales": total}
        for date_obj, total in daily_totals.items()
    ]
    daily_records.sort(key=lambda x: x["date"])

    return daily_records

def compute_monthly_summary(daily_records):
    """Compute month-by-month sales totals for the current year, January through the current month.

    Only records whose year matches today's year are included, so the output
    automatically rolls over to the new year's months once the calendar year changes.

    Args:
        daily_records: List of dicts with 'date' (datetime.date) and 'daily_sales'

    Returns:
        List of dicts ordered January -> current month:
        [{"month": "January", "total_sales": 45230}, ...]
    """
    today = datetime.now(ZoneInfo("Asia/Kolkata")).date()

    monthly_totals = {month_num: 0 for month_num in range(1, today.month + 1)}
    for record in daily_records:
        date = record["date"]
        if date.year != today.year or date.month > today.month:
            continue
        monthly_totals[date.month] += record["daily_sales"]

    monthly_summary = []
    for month_num in range(1, today.month + 1):
        month_name = datetime(today.year, month_num, 1).strftime("%B")
        monthly_summary.append({
            "month": month_name,
            "total_sales": int(monthly_totals[month_num]),
        })

    return monthly_summary

def format_ist_timestamp():
    """Return the current IST time formatted as 'D Month YYYY, H:MM AM/PM IST'."""
    now_ist = datetime.now(ZoneInfo("Asia/Kolkata"))
    hour_12 = now_ist.strftime("%I").lstrip("0") or "12"
    return f"{now_ist.day} {now_ist.strftime('%B %Y')}, {hour_12}:{now_ist.strftime('%M %p')} IST"

def write_monthly_json(year, monthly_summary, output_file):
    """Write year-scoped monthly summary to JSON file.

    Args:
        year: int, the year the monthly_summary belongs to
        monthly_summary: List of dicts with 'month' and 'total_sales'
        output_file: Path to output JSON file
    """
    monthly_data = {
        "year": year,
        "monthly_chart_data": monthly_summary,
        "last_updated": format_ist_timestamp(),
    }

    with open(output_file, 'w') as f:
        json.dump(monthly_data, f, indent=2)

    print(f"Monthly dashboard data written to {output_file}")

def main():
    """Main entry point for the monthly fetch and compute script."""
    try:
        # Step 1: Fetch Excel from SharePoint via Graph API
        print("Fetching Daily_Invoice.xlsx from SharePoint...")
        wb = fetch_excel_from_sharepoint()

        # Step 2: Validate sheet exists
        if "invoice_transactions" not in wb.sheetnames:
            print("ERROR: Sheet 'invoice_transactions' not found in Excel file")
            exit(1)

        ws = wb["invoice_transactions"]
        print("[OK] Successfully fetched and opened Daily_Invoice.xlsx")

        # Step 3: Parse and aggregate invoice_transactions by date
        daily_records = parse_excel(ws)

        if not daily_records:
            print("WARNING: No valid data rows found in Excel file")
            daily_records = []
        else:
            print(f"Parsed {len(daily_records)} daily records from Excel")
            print(f"Date range: {daily_records[0]['date']} to {daily_records[-1]['date']}")

        # Step 4: Compute month-by-month summary for the current year
        today = datetime.now(ZoneInfo("Asia/Kolkata")).date()
        monthly_summary = compute_monthly_summary(daily_records)
        print(f"Computed {len(monthly_summary)} monthly summaries for {today.year}")
        for entry in monthly_summary:
            print(f"  {entry['month']}: {entry['total_sales']}")

        # Step 5: Write output JSON
        write_monthly_json(today.year, monthly_summary, OUTPUT_FILE)

        print("SUCCESS: Monthly dashboard data computed and written")

        wb.close()

    except RuntimeError as e:
        print(f"ERROR: {e}")
        exit(1)
    except Exception as e:
        print(f"ERROR: Unexpected error: {e}")
        exit(1)

if __name__ == "__main__":
    main()
